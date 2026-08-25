"""Read-only audit for the UCI Online Retail workbook.

This module deliberately produces audit metadata only. It does not emit cleaned
records, sampled items, ground truth, privacy-mechanism output, or MSE values.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


REQUIRED_COLUMNS = [
    "InvoiceNo", "StockCode", "Description", "Quantity",
    "InvoiceDate", "UnitPrice", "CustomerID", "Country",
]
STOCK_CODE_RE = re.compile(r"^\d{5}[A-Z]{0,2}$")
PRIMARY_REASON_PRIORITY = [
    "invalid_customer_id",
    "cancelled_invoice",
    "nonpositive_quantity",
    "nonpositive_unit_price",
    "invalid_stock_code",
]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _json_value(value: Any) -> Any:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.isoformat()
    if hasattr(value, "item"):
        return value.item()
    return value


def normalize_customer_id(series: pd.Series) -> tuple[pd.Series, pd.Series]:
    numeric = pd.to_numeric(series, errors="coerce")
    finite = numeric.notna() & numeric.map(lambda value: math.isfinite(float(value)) if pd.notna(value) else False)
    integral = finite & (numeric % 1 == 0)
    normalized = numeric.where(integral).astype("Int64")
    return normalized, integral


def normalize_stock_code(series: pd.Series) -> pd.Series:
    return series.map(lambda value: "" if pd.isna(value) else str(value).strip().upper())


def _quantiles(lengths: pd.Series) -> dict[str, float | int | None]:
    if lengths.empty:
        return {key: None for key in ("min", "p25", "median", "p75", "p90", "p95", "p99", "max", "mean")}
    values = lengths.astype(float)
    return {
        "min": int(values.min()), "p25": float(values.quantile(.25)),
        "median": float(values.quantile(.5)), "p75": float(values.quantile(.75)),
        "p90": float(values.quantile(.9)), "p95": float(values.quantile(.95)),
        "p99": float(values.quantile(.99)), "max": int(values.max()),
        "mean": float(values.mean()),
    }


def audit_workbook(path: Path, *, source: dict[str, Any] | None = None) -> dict[str, Any]:
    path = path.resolve()
    try:
        report_path = path.relative_to(Path.cwd().resolve()).as_posix()
    except ValueError:
        report_path = str(path)
    before_stat = path.stat()
    before_hash = sha256_file(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = workbook.sheetnames
    workbook.close()
    if len(sheets) != 1:
        raise ValueError(f"Expected exactly one worksheet, found {sheets}")

    frame = pd.read_excel(path, sheet_name=sheets[0], engine="openpyxl")
    missing_columns = [column for column in REQUIRED_COLUMNS if column not in frame.columns]
    extra_columns = [str(column) for column in frame.columns if column not in REQUIRED_COLUMNS]
    if missing_columns:
        raise ValueError(f"Missing required columns: {missing_columns}")

    customer_id, valid_customer = normalize_customer_id(frame["CustomerID"])
    invoice = frame["InvoiceNo"].map(lambda value: "" if pd.isna(value) else str(value).strip())
    cancelled = invoice.str.upper().str.startswith("C")
    quantity = pd.to_numeric(frame["Quantity"], errors="coerce")
    unit_price = pd.to_numeric(frame["UnitPrice"], errors="coerce")
    stock = normalize_stock_code(frame["StockCode"])
    valid_stock = stock.str.fullmatch(STOCK_CODE_RE)

    rule_masks = {
        "invalid_customer_id": ~valid_customer,
        "cancelled_invoice": cancelled,
        "nonpositive_quantity": quantity.isna() | (quantity <= 0),
        "nonpositive_unit_price": unit_price.isna() | (unit_price <= 0),
        "invalid_stock_code": ~valid_stock,
    }
    primary = pd.Series("kept", index=frame.index, dtype="object")
    unassigned = pd.Series(True, index=frame.index)
    for reason in PRIMARY_REASON_PRIORITY:
        assigned = unassigned & rule_masks[reason]
        primary.loc[assigned] = reason
        unassigned &= ~assigned
    kept = primary.eq("kept")

    kept_pairs = pd.DataFrame({"customer_id": customer_id[kept], "stock_code": stock[kept]})
    duplicate_pair_mask = kept_pairs.duplicated(["customer_id", "stock_code"], keep="first")
    baskets = kept_pairs.drop_duplicates().groupby("customer_id")["stock_code"].agg(list)
    basket_lengths = baskets.map(len)
    mapped_codes = sorted(set(kept_pairs["stock_code"]))
    mapping_preview = [
        {"stock_code": code, "internal_id": index}
        for index, code in enumerate(mapped_codes[:10], start=1)
    ]

    numeric_raw = sorted({int(code) for code in stock[stock.str.fullmatch(r"\d+")]})
    numeric_contiguous = bool(numeric_raw) and numeric_raw == list(range(numeric_raw[0], numeric_raw[-1] + 1))
    valid_customers_before = set(customer_id[valid_customer].astype(int).tolist())
    valid_customers_after = set(baskets.index.astype(int).tolist())
    empty_after_filter = valid_customers_before - valid_customers_after

    dtype_report = {
        column: {
            "pandas_dtype": str(frame[column].dtype),
            "python_types": dict(Counter(type(value).__name__ for value in frame[column].dropna())),
        }
        for column in REQUIRED_COLUMNS
    }
    missing = {column: int(frame[column].isna().sum()) for column in REQUIRED_COLUMNS}
    primary_counts = {reason: int((primary == reason).sum()) for reason in PRIMARY_REASON_PRIORITY}
    primary_counts["kept"] = int(kept.sum())

    after_stat = path.stat()
    after_hash = sha256_file(path)
    unchanged = before_hash == after_hash and before_stat.st_size == after_stat.st_size and before_stat.st_mtime_ns == after_stat.st_mtime_ns
    if not unchanged:
        raise RuntimeError("Raw workbook changed during the audit")

    report = {
        "schema_version": 1,
        "audit_scope": "read_only_data_audit_only_no_sampling_no_gt_no_mechanisms_no_mse",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "source": source or {},
        "file": {
            "name": path.name, "path": report_path, "size_bytes": before_stat.st_size,
            "sha256": before_hash, "unchanged_during_audit": unchanged,
        },
        "workbook": {"worksheets": sheets, "active_worksheet": sheets[0]},
        "schema": {
            "columns": [str(column) for column in frame.columns],
            "required_columns": REQUIRED_COLUMNS, "missing_columns": missing_columns,
            "extra_columns": extra_columns, "dtypes": dtype_report,
        },
        "raw": {
            "detail_rows": int(len(frame)), "missing_values": missing,
            "fully_duplicated_detail_rows": int(frame.duplicated().sum()),
            "unique_customer_ids_with_valid_integer_form": len(valid_customers_before),
            "unique_normalized_stock_codes": int(stock.nunique()),
            "noninteger_or_missing_customer_id_rows": int((~valid_customer).sum()),
            "noncanonical_stock_code_rows": int((stock != frame["StockCode"].map(lambda x: "" if pd.isna(x) else str(x))).sum()),
            "numeric_stock_code_domain": {
                "unique_count": len(numeric_raw),
                "minimum": numeric_raw[0] if numeric_raw else None,
                "maximum": numeric_raw[-1] if numeric_raw else None,
                "starts_at_1": bool(numeric_raw) and numeric_raw[0] == 1,
                "is_contiguous": numeric_contiguous,
                "missing_integer_count_between_min_and_max": (numeric_raw[-1] - numeric_raw[0] + 1 - len(numeric_raw)) if numeric_raw else 0,
            },
        },
        "filtering": {
            "priority": PRIMARY_REASON_PRIORITY,
            "independent_rule_hit_counts": {key: int(mask.sum()) for key, mask in rule_masks.items()},
            "unique_primary_reason_counts": primary_counts,
            "accounting_total": sum(primary_counts.values()),
            "all_rows_accounted_for": sum(primary_counts.values()) == len(frame),
        },
        "aggregation_preview": {
            "definition": "One user per basket after filtering; StockCode is normalized and deduplicated within each user.",
            "users_before_row_filters": len(valid_customers_before),
            "users_after_row_filters": len(valid_customers_after),
            "items_before_row_filters": int(stock[valid_stock].nunique()),
            "items_after_row_filters": len(mapped_codes),
            "empty_baskets_after_filtering": len(empty_after_filter),
            "kept_detail_rows": int(kept.sum()),
            "duplicate_user_item_rows_removed_by_basket_deduplication": int(duplicate_pair_mask.sum()),
            "users_with_duplicate_items": int(kept_pairs.loc[kept_pairs.duplicated(["customer_id", "stock_code"], keep=False), "customer_id"].nunique()),
            "basket_length": _quantiles(basket_lengths),
            "all_users_unique": bool(baskets.index.is_unique),
            "all_baskets_nonempty": bool((basket_lengths > 0).all()),
            "all_baskets_deduplicated": all(len(items) == len(set(items)) for items in baskets),
        },
        "mapping_preview": {
            "rule": "Sort eligible normalized StockCode values lexicographically and assign consecutive IDs 1..d.",
            "domain_size_d": len(mapped_codes), "first_10_forward_entries": mapping_preview,
            "reverse_mapping_is_defined": True, "mapping_materialized": False,
        },
        "prohibited_outputs": {"fixed_single_item_dataset": False, "ground_truth": False, "mechanism_outputs": False, "mse": False},
    }
    return report


def markdown_summary(report: dict[str, Any]) -> str:
    raw = report["raw"]
    filtering = report["filtering"]
    agg = report["aggregation_preview"]
    lines = [
        "# UCI Online Retail 数据只读审计摘要", "",
        f"- 原始文件：`{report['file']['name']}`（{report['file']['size_bytes']:,} bytes）",
        f"- SHA-256：`{report['file']['sha256']}`",
        f"- 工作表：`{report['workbook']['active_worksheet']}`；交易明细行：{raw['detail_rows']:,}",
        f"- 过滤后交易明细：{agg['kept_detail_rows']:,}；用户：{agg['users_after_row_filters']:,}；商品：{agg['items_after_row_filters']:,}",
        f"- 用户内去重移除的重复商品明细：{agg['duplicate_user_item_rows_removed_by_basket_deduplication']:,}",
        f"- 篮长：min={agg['basket_length']['min']}，median={agg['basket_length']['median']}，p95={agg['basket_length']['p95']}，max={agg['basket_length']['max']}",
        "", "## 唯一主排除原因（固定优先级）", "",
    ]
    for reason in PRIMARY_REASON_PRIORITY:
        lines.append(f"- `{reason}`：{filtering['unique_primary_reason_counts'][reason]:,}")
    lines.extend([
        "", "所有原始行均已唯一归入“保留”或一个主排除原因。原始 Excel 是一行一个交易明细，只有按 `CustomerID` 聚合并在用户内对 `StockCode` 去重后，才形成“一行一个用户、变长购物篮”的逻辑数据。", "",
        "本阶段仅预览字典序稳定映射 `StockCode -> 1..d`，未生成映射文件、固定单商品数据、GT、隐私机制输出或 MSE。", "",
    ])
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path)
    parser.add_argument("--json", type=Path, default=Path("results/data_audit.json"))
    parser.add_argument("--markdown", type=Path, default=Path("results/data_audit.md"))
    args = parser.parse_args()
    source = {
        "dataset": "UCI Online Retail", "uci_dataset_id": 352,
        "dataset_page": "https://archive.ics.uci.edu/dataset/352/online%2Bretail",
        "download_url": "https://archive.ics.uci.edu/static/public/352/online+retail.zip",
        "doi": "10.24432/C5BW33", "license": "CC BY 4.0",
    }
    metadata_path = args.input.parent / "source_metadata.json"
    if metadata_path.exists():
        source.update(json.loads(metadata_path.read_text(encoding="utf-8")))
    report = audit_workbook(args.input, source=source)
    args.json.parent.mkdir(parents=True, exist_ok=True)
    args.markdown.parent.mkdir(parents=True, exist_ok=True)
    args.json.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    args.markdown.write_text(markdown_summary(report), encoding="utf-8")


if __name__ == "__main__":
    main()
